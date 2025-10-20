import logging
import math
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import numpy as np
import pandas as pd
import os
from typing import Dict, List, Optional, Tuple
from itertools import combinations, groupby, product
from datetime import datetime
import heapq
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side
from data_models import CalculationParams, CalculationResult, ProcessingMode, TotoResult




def calculation_to_toto_result(
    calculation_result: CalculationResult, mode: ProcessingMode
) -> TotoResult:
    sikum = None
    if all([p[1] == p[0] for p in calculation_result.chosen_columns_ranges]):
        sikum = sum([p[0] for p in calculation_result.chosen_columns_ranges])
    return TotoResult(
        turim=[c + 1 for c in calculation_result.chosen_columns],
        mahzorim=[r + 1 for r in calculation_result.chosen_rows],
        tvachim=[
            f"{start}-{end}" if start != end else str(start)
            for start, end in calculation_result.chosen_columns_ranges
        ]
        * (
            len(calculation_result.chosen_columns)
            if mode == ProcessingMode.AGGREGATIVE
            else 1
        ),
        sikum=sikum,
    )




class MatrixApp:
    def __init__(self, root):
        self.root = root
        self.root.title("מחשבון קומבינציות")
        self.root.geometry("500x600")
        self.root.configure(bg="#f0f0f0")

        # Configure style
        style = ttk.Style()
        style.theme_use("clam")

        self.file_path = None

        # Main container
        main_frame = tk.Frame(root, bg="#f0f0f0", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Title
        title_label = tk.Label(
            main_frame,
            text="מחשבון קומבינציות",
            font=("Arial", 16, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
        )
        title_label.pack(pady=(0, 20))

        # File selection frame
        file_frame = tk.LabelFrame(
            main_frame,
            text="בחירת קובץ",
            font=("Arial", 10, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=15,
            pady=10,
        )
        file_frame.pack(fill="x", pady=(0, 15))

        self.file_label = tk.Label(
            file_frame,
            text="לא נבחר קובץ",
            bg="#f0f0f0",
            fg="#7f8c8d",
            font=("Arial", 9),
        )
        self.file_label.pack(pady=(0, 5))

        self.browse_button = tk.Button(
            file_frame,
            text="בחר קובץ Excel",
            command=self.browse_file,
            bg="#3498db",
            fg="white",
            font=("Arial", 9, "bold"),
            relief="flat",
            padx=15,
            pady=5,
        )
        self.browse_button.pack()

        # Parameters frame
        params_frame = tk.LabelFrame(
            main_frame,
            text="פרמטרים",
            font=("Arial", 10, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50",
            padx=15,
            pady=10,
        )
        params_frame.pack(fill="x", pady=(0, 15))

        # Number of configurations (N)
        n_frame = tk.Frame(params_frame, bg="#f0f0f0")
        n_frame.pack(fill="x", pady=(0, 10))

        self.n_label = tk.Label(
            n_frame, text="מספר טורים רצוי", bg="#f0f0f0", font=("Arial", 9)
        )
        self.n_label.pack(anchor="w")

        self.n_var = tk.IntVar(value=2)
        self.n_entry = tk.Entry(
            n_frame, textvariable=self.n_var, font=("Arial", 9), width=10
        )
        self.n_entry.pack(anchor="w", pady=(5, 0))

        # Top K configurations
        topk_frame = tk.Frame(params_frame, bg="#f0f0f0")
        topk_frame.pack(fill="x", pady=(0, 10))

        self.topk_label = tk.Label(
            topk_frame,
            text="מספר תצורות להצגה (טופ K)",
            bg="#f0f0f0",
            font=("Arial", 9),
        )
        self.topk_label.pack(anchor="w")

        self.topk_var = tk.IntVar(value=1)
        self.topk_entry = tk.Entry(
            topk_frame, textvariable=self.topk_var, font=("Arial", 9), width=10
        )
        self.topk_entry.pack(anchor="w", pady=(5, 0))
        # self.n_entry.bind('<KeyRelease>', self.on_n_change)

        # Relaxation number
        relax_frame = tk.Frame(params_frame, bg="#f0f0f0")
        relax_frame.pack(fill="x", pady=(0, 10))
        self.relax_label = tk.Label(
            relax_frame, text="מספר קפיצות מקסימלי", bg="#f0f0f0", font=("Arial", 9)
        )
        self.relax_label.pack(anchor="w")
        self.relax_var = tk.IntVar(value=0)
        self.relax_entry = tk.Entry(
            relax_frame, textvariable=self.relax_var, font=("Arial", 9), width=10
        )
        self.relax_entry.pack(anchor="w", pady=(5, 0))

        # Processing mode
        mode_frame = tk.Frame(params_frame, bg="#f0f0f0")
        mode_frame.pack(fill="x", pady=(0, 10))

        self.mode_label = tk.Label(
            mode_frame, text="מצב עיבוד", bg="#f0f0f0", font=("Arial", 9)
        )
        self.mode_label.pack(anchor="w")

        self.mode_var = tk.StringVar(value=ProcessingMode.SEPARATE.value)
        self.agg_radio = tk.Radiobutton(
            mode_frame,
            text="סיכומי",
            variable=self.mode_var,
            value=ProcessingMode.AGGREGATIVE.value,
            bg="#f0f0f0",
            font=("Arial", 9),
        )
        self.sep_radio = tk.Radiobutton(
            mode_frame,
            text="נפרד",
            variable=self.mode_var,
            value=ProcessingMode.SEPARATE.value,
            bg="#f0f0f0",
            font=("Arial", 9),
        )
        self.agg_radio.pack(anchor="w", pady=(5, 0))
        self.sep_radio.pack(anchor="w")

        # Run button
        self.run_button = tk.Button(
            main_frame,
            text="הרץ חישוב",
            command=self.run,
            bg="#27ae60",
            fg="white",
            font=("Arial", 11, "bold"),
            relief="flat",
            padx=30,
            pady=10,
        )
        self.run_button.pack(pady=10)

    def browse_file(self):
        filetypes = [("Excel files", "*.xlsx *.xls")]
        path = filedialog.askopenfilename(
            title="Select Excel File", filetypes=filetypes
        )
        if path:
            self.file_path = path
            self.file_label.config(text=os.path.basename(path), fg="#2c3e50")

    def on_n_change(self, event=None):
        try:
            n = int(self.n_entry.get())
            if n < 1:
                n = 1
        except ValueError:
            n = 1
        self.n_var.set(n)

    @staticmethod
    def load_matrix_from_excel(file_path) -> np.ndarray:
        df = pd.read_excel(file_path, header=None)
        c = df.isna().iloc[:, 0]
        r = df.isna().iloc[0]
        c_min = r[r].index[0] if sum(r) > 0 else None
        r_min = c[c].index[0] if sum(c) > 0 else None
        df = df.iloc[:r_min, :c_min].dropna().astype(int)
        return df.values

    @staticmethod
    def filter_columns_by_ranges(matrix, ranges):
        filtered_columns = []
        for col_idx, (min_val, max_val) in enumerate(ranges):
            col = matrix[:, col_idx]
            indices = {
                row_idx for row_idx, x in enumerate(col) if min_val <= x <= max_val
            }
            filtered_columns.append(indices)
        return filtered_columns

    @staticmethod
    def max_n_representative_intersection(
        groups, n, top_k, stats: Optional[dict], enable_pruning: bool = False
    ) -> List[Tuple[list[int], list[int]]]:
        """
        groups: O(C*R)
        """

        heap = []
        if stats is None:
            stats = {}

        stats["combinations"] = len(groups) ** n
        stats["total_choices"] = 0
        stats["max_choices_per_combination"] = 0
        stats["pruned_choices"] = 0
        stats["pruned_combinations"] = 0
        max_choices_time = 0.0
        
        # Track minimum threshold for pruning (only used when pruning is enabled)
        min_threshold = 0 if enable_pruning else -1

        # All combinations of n groups out of N
        t0 = time.perf_counter()
        for group_indices in combinations(range(len(groups)), n):  # O(C^n)
            selected_groups = [groups[i] for i in group_indices]
            # All ways to pick one subset from each selected group

            num_choices = math.prod(len(g) for g in selected_groups)
            stats["total_choices"] += num_choices
            stats["max_choices_per_combination"] = max(
                stats["max_choices_per_combination"], num_choices
            )
            t2 = time.perf_counter()
            choices_pruned = 0
            
            for choice in product(
                *[range(len(g)) for g in selected_groups]
            ):  # O(max(groups(R))^n)
                selected_subsets = [selected_groups[i][choice[i]] for i in range(n)]
                current_intersection = selected_subsets[0]
                
                # Early pruning based on initial set size (only when pruning enabled)
                if enable_pruning and len(current_intersection) < min_threshold:
                    choices_pruned += 1
                    stats["pruned_choices"] += 1
                    continue
                
                # Compute intersection (with optional early termination)
                pruned = False
                for subset in selected_subsets[1:]:
                    current_intersection = current_intersection.intersection(subset)
                    
                    # Early termination if intersection becomes too small (only when pruning enabled)
                    if enable_pruning and len(current_intersection) < min_threshold:
                        choices_pruned += 1
                        stats["pruned_choices"] += 1
                        pruned = True
                        break
                
                if pruned:
                    continue

                score = len(current_intersection)
                
                # Only consider adding if score is good enough (or pruning disabled)
                if not enable_pruning or score >= min_threshold:
                    # Use a min-heap of size top_k
                    item = (score, group_indices, sorted(current_intersection))
                    if len(heap) < top_k:
                        heapq.heappush(heap, item)
                        # Update threshold when heap becomes full (only when pruning enabled)
                        if enable_pruning and len(heap) == top_k:
                            min_threshold = heap[0][0]
                    else:
                        if item > heap[0]:
                            heapq.heappushpop(heap, item)
                            # Update threshold with new minimum (only when pruning enabled)
                            if enable_pruning:
                                min_threshold = heap[0][0]
            
            if choices_pruned == num_choices:
                stats["pruned_combinations"] += 1
                
            t3 = time.perf_counter()
            max_choices_time = max(max_choices_time, t3 - t2)
        t1 = time.perf_counter()
        stats["combinations_time_s"] = t1 - t0
        stats["max_choices_time"] = max_choices_time
        stats["average_choices_per_combination"] = (
            stats["total_choices"] / stats["combinations"]
        )
        if enable_pruning and stats["total_choices"] > 0:
            stats["pruning_rate"] = stats["pruned_choices"] / stats["total_choices"]

        # Sort results by intersection size descending, then by group indices
        results = sorted(heap, key=lambda x: (x[0], x[1]), reverse=True)
        return [
            (list(group_indices), intersection)
            for _, group_indices, intersection in results
        ]

    @staticmethod
    def max_n_subset_intersection_brute_force(
        matrix: np.ndarray, params: CalculationParams, stats: Optional[Dict]
    ) -> List[CalculationResult]:
        per_col_subsets = []

        for col in matrix.T:  # O(C)
            index_and_elem_pairs = sorted(
                list(enumerate(col)), key=lambda p: p[1]
            )  # O(RlogR)
            groups = [
                list(g) for _, g in groupby(index_and_elem_pairs, key=lambda p: p[1])
            ]  # O(R)
            if params.relaxation > 0:
                relaxed_groups = []
                for group_index, group in enumerate(groups):
                    relaxed_group = group.copy()
                    relaxed_groups.append(relaxed_group)
                    if len(group) == 0:
                        continue
                    group_value = group[0][1]
                    for added_group in groups[group_index + 1 :]:
                        if (added_group[0][1] - group_value) > params.relaxation:
                            break
                        relaxed_group.extend(added_group.copy())
                groups = relaxed_groups
            subsets = [set([p[0] for p in group]) for group in groups]
            per_col_subsets.append(subsets)

        top_intersections = MatrixApp.max_n_representative_intersection(
            per_col_subsets, n=params.n, top_k=params.top_k, stats=stats,
            enable_pruning=params.enable_pruning
        )

        res = []
        for intersection in top_intersections:
            chosen_columns, chosen_rows = intersection
            chosen_columns_ranges = []

            for col in chosen_columns:
                view = matrix[chosen_rows, col]
                col_range = (view.min(), view.max())
                chosen_columns_ranges.append(col_range)
                assert col_range[1] - col_range[0] <= params.relaxation, (
                    "invalid relaxation"
                )
            res.append(
                CalculationResult(
                    chosen_columns=chosen_columns,
                    chosen_rows=chosen_rows,
                    chosen_columns_ranges=chosen_columns_ranges,
                )
            )

        return res

    @staticmethod
    def max_n_subset_sums_brute_force(
        matrix: np.ndarray, params: CalculationParams, stats: Optional[Dict]
    ) -> List[CalculationResult]:
        if params.relaxation > 0:
            raise NotImplementedError("כרגע אין תמיכה בקפיצות במצב חישוב סיכומי")
        n = params.n
        top_k = params.top_k
        heap = []


        stats["num_combinations"] = matrix.shape[1] ** n

        t0 = time.perf_counter()
        for combo_indices in combinations(range(matrix.shape[1]), n):
            sums: np.ndarray = matrix[:, combo_indices].sum(axis=1)
            counts = np.bincount(sums)
            max_s = np.argmax(counts)
            max_count = counts[max_s]
            chosen_rows = sorted(list(np.nonzero((sums == max_s)))[0])
            chosen_rows_range = [(max_s, max_s)]
            item = (max_count, combo_indices, chosen_rows, chosen_rows_range)
            if len(heap) < top_k:
                heapq.heappush(heap, item)
            else:
                if item > heap[0]:
                    heapq.heappushpop(heap, item)

        # Sort results by count descending, then by columns
        results = sorted(heap, key=lambda x: (x[0], x[1]), reverse=True)

        t1 = time.perf_counter()
        stats["total_time_elapsed_s"] = t1 - t0
        return [
            CalculationResult(
                chosen_columns=list(combo_indices),
                chosen_rows=chosen_rows,
                chosen_columns_ranges=chosen_rows_range,
            )
            for _, combo_indices, chosen_rows, chosen_rows_range in results
        ]

    @staticmethod
    def run_calculation(
        matrix, params: CalculationParams, return_stats: bool = False
    ) -> Tuple[List[CalculationResult], Optional[Dict]]:
        if params.mode == ProcessingMode.SEPARATE:
            calculation_func = MatrixApp.max_n_subset_intersection_brute_force
        elif params.mode == ProcessingMode.AGGREGATIVE:
            calculation_func = MatrixApp.max_n_subset_sums_brute_force

        stats = {} if return_stats else None
        res = calculation_func(matrix, params, stats)
        return res, stats

    def run(self):
        if not self.file_path:
            messagebox.showerror("Error", "בחר קובץ אקסל")
            return
        try:
            matrix = self.load_matrix_from_excel(self.file_path)
            n = self.n_var.get()
            top_k = self.topk_var.get()
            mode_str = self.mode_var.get()
            try:
                mode = ProcessingMode(mode_str)
            except ValueError:
                messagebox.showerror("Error", f"Invalid processing mode: {mode_str}")
                return
            relaxation = self.relax_var.get()
            params = CalculationParams(
                n=n, mode=mode, relaxation=relaxation, top_k=top_k
            )
            results, _ = self.run_calculation(matrix, params)

            toto_results = [
                calculation_to_toto_result(res, params.mode) for res in results
            ]

            topk_colname = f"טופ {params.top_k}"

            metadata_df = pd.DataFrame(
                data={
                    "כמות עמודות": [params.n],
                    "מצב חישוב": [params.mode.value],
                    "מספר קפיצות מקסימלי": [params.relaxation],
                    "כמות מחזורים מקסימלית": [
                        len(toto_results[-1].mahzorim) if toto_results else 0
                    ],
                }
            )

            data_df = pd.DataFrame(
                data={
                    "מספר פעמים": [len(result.mahzorim) for result in toto_results],
                    topk_colname: [i + 1 for i in range(params.top_k)],
                    "עמודות": [result.turim for result in toto_results],
                    "טווחים": [result.tvachim for result in toto_results],
                    "מחזורים": [result.mahzorim for result in toto_results],
                    "סיכום": [result.sikum for result in toto_results],
                }
            )

            # Change output file extension to xlsx
            base_input_path, _ = os.path.splitext(self.file_path)
            time_addition = datetime.strftime(datetime.now(), format="%Y%m%d_%H%M%S")
            base_output_path = base_input_path + f"_output_{time_addition}"
            default_output_path = base_output_path + ".xlsx"
            output_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="שמירת תוצאה",
                initialfile=default_output_path,
            )
            if output_path:
                # Create Excel file with openpyxl engine
                with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                    # If processing mode is SEPARATE, remove 'סיכום' column
                    if params.mode == ProcessingMode.SEPARATE:
                        if "סיכום" in data_df.columns:
                            data_df = data_df.drop(columns=["סיכום"])

                    # Write metadata first
                    metadata_df.to_excel(
                        writer, sheet_name="Sheet1", index=False, startrow=0, startcol=0
                    )

                    # Get the workbook and active sheet
                    worksheet = writer.sheets["Sheet1"]

                    # Style the metadata table with red background and bold font
                    red_fill = PatternFill(
                        start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
                    )
                    bold_font = Font(bold=True)

                    # Apply styling to metadata table (headers and data)
                    for row in range(1, metadata_df.shape[0] + 2):
                        for col in range(1, metadata_df.shape[1] + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.fill = red_fill
                            cell.font = bold_font

                    # Calculate starting row for data table (metadata rows + 2 blank rows)
                    data_start_row = (metadata_df.shape[0] + 1) + 4

                    value_cols = ["מספר פעמים", topk_colname] + (
                        ["סיכום"] if params.mode == ProcessingMode.AGGREGATIVE else []
                    )
                    list_cols = ["עמודות", "מחזורים"]

                    num_entries = len(data_df)
                    last_available_col = 1
                    space_factor = 3

                    def set_cell_value(cell, value):
                        # This function sets the value of a cell and can be extended
                        # to add more functionality in the future
                        cell.value = value
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        thin_border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )
                        cell.border = thin_border

                    for value_col in value_cols:
                        header_cell = worksheet.cell(
                            row=data_start_row - 1, column=last_available_col
                        )
                        set_cell_value(header_cell, value_col)
                        if value_col == "מספר פעמים":
                            for row in range(num_entries):
                                cell = worksheet.cell(
                                    row=data_start_row + space_factor * row,
                                    column=last_available_col,
                                )
                                set_cell_value(cell, data_df.at[row, "מספר פעמים"])
                        elif value_col == topk_colname:
                            for row in range(num_entries):
                                cell = worksheet.cell(
                                    row=data_start_row + space_factor * row,
                                    column=last_available_col,
                                )
                                set_cell_value(cell, data_df.at[row, topk_colname])
                        elif value_col == "סיכום":
                            for row in range(num_entries):
                                cell = worksheet.cell(
                                    row=data_start_row + space_factor * row,
                                    column=last_available_col,
                                )
                                set_cell_value(cell, data_df.at[row, "סיכום"])
                        last_available_col += 1

                    for list_col in list_cols:
                        col_width = data_df[list_col].map(len).max()
                        header_cell = worksheet.cell(
                            row=data_start_row - 1, column=last_available_col
                        )
                        set_cell_value(header_cell, list_col)
                        worksheet.merge_cells(
                            start_row=data_start_row - 1,
                            start_column=last_available_col,
                            end_row=data_start_row - 1,
                            end_column=last_available_col + col_width - 1,
                        )

                        if list_col == "עמודות":
                            for row in range(num_entries):
                                for col_offset, amuda_and_tvach in enumerate(
                                    zip(
                                        data_df.at[row, "עמודות"],
                                        data_df.at[row, "טווחים"],
                                    )
                                ):
                                    amuda, tvach = amuda_and_tvach
                                    cell = worksheet.cell(
                                        row=data_start_row + space_factor * row,
                                        column=last_available_col + col_offset,
                                    )
                                    set_cell_value(cell, amuda)
                                    cell = worksheet.cell(
                                        row=data_start_row + space_factor * row + 1,
                                        column=last_available_col + col_offset,
                                    )
                                    set_cell_value(cell, tvach)
                        elif list_col == "מחזורים":
                            for row in range(num_entries):
                                for col_offset, machzor in enumerate(
                                    data_df.at[row, "מחזורים"]
                                ):
                                    cell = worksheet.cell(
                                        row=data_start_row + space_factor * row,
                                        column=last_available_col + col_offset,
                                    )
                                    set_cell_value(cell, machzor)
                        last_available_col += col_width
                messagebox.showinfo("Success", f"Result saved to {output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process file: {e}")


def main():
    root = tk.Tk()
    app = MatrixApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
